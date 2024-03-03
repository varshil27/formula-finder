import openpyxl
import json
from algo import get_all_formulas
from get_data import read_excel
def save_excel(file_path, row_skip, col_skip,formula_signs):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    max_row = sheet.max_row
    max_col = sheet.max_column

    for col_sheet in range(col_skip + 1, max_col + 1):
        col = chr(ord('A')+col_sheet-1)
        for row_sheet in range(row_skip + 1, max_row + 1):
            row = row_sheet-(row_skip+1)
            if len(formula_signs[row]) > 0:
                generated_formula = "="
                for i in formula_signs[row]:
                    if i > 0:
                        generated_formula += " + "
                    else :
                        generated_formula +=" - "
                    generated_formula += col
                    generated_formula += str(abs(i)+row_skip)
                cell = sheet[col + str(row_sheet)]
                cell.value = generated_formula

    print("""File sucessfully saved!!""")
    workbook.save('output.xlsx')
    workbook.close()

if __name__ == "__main__":
    with open("config.json", "r") as config_file:
        config = json.load(config_file)
    data = read_excel(config["file_path"], config["row_skip"], config["col_skip"])
    formula_signs = get_all_formulas(data)
    save_excel(config["file_path"], config["row_skip"], config["col_skip"],formula_signs)