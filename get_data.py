import openpyxl
import json
def convert_to_int(value):
    """Converts a value to an integer.
    Parameters:
        - value (str or int): The value to be converted to an integer.
    Returns:
        - int: The converted integer value.
    Processing Logic:
        - Handles None values.
        - Returns the value unchanged if it's already an integer.
        - Handles negative values in parentheses.
        - Handles comma-separated values."""
    if value is None:
        return 0
    elif isinstance(value, int):
        return value  # Return the value unchanged if it's already an integer
    elif "(" in value and ")" in value:
        return -int(value[1:-1].replace(',', ''))
    else:
        return int(value.replace(',', ''))

def read_excel(file_path, row_skip, col_skip):

    """Reads an excel file and returns all columns' values as a list of lists.
    Parameters:
        - file_path (str): Path to the excel file.
        - row_skip (int): Number of rows to skip from the top.
        - col_skip (int): Number of columns to skip from the left.
    Returns:
        - list: List of lists containing all columns' values.
    Processing Logic:
        - Load excel file using openpyxl.
        - Get active sheet.
        - Get max row and column values.
        - Loop through columns and extract values.
        - Convert values to integers.
        - Append column values to a list.
        - Close workbook.
        - Return list of lists containing all columns' values."""
    workbook = openpyxl.load_workbook(file_path)


    sheet = workbook.active
    all_columns_values = []

    max_row = sheet.max_row
    max_col = sheet.max_column


    for col in range(col_skip + 1, max_col + 1):

        column_values = [sheet.cell(row=row, column=col).value for row in range(row_skip + 1, max_row + 1)]
        column_values = [convert_to_int(item) for item in column_values]
        all_columns_values.append(column_values)

    workbook.close()
    return all_columns_values

if __name__ == "__main__":
    with open("config.json", "r") as config_file:
        config = json.load(config_file)

    read_excel(config["file_path"], config["row_skip"], config["col_skip"])